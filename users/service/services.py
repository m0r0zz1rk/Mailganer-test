from openpyxl import load_workbook
from rest_framework.generics import get_object_or_404

from users.models import Subscribers, SubLists


def ImportSubsFromList(id: int, file) -> bool:
    """Чтение данных и добавление новых подписчиков"""
    list = get_object_or_404(queryset=SubLists.objects.all(), pk=id)
    book = load_workbook(file)
    try:
        for name_sh in book.sheetnames:
            sheet = book[name_sh]
            for row in range(1, sheet.max_row+1):
                new_sub = Subscribers(
                    email=sheet.cell(row=row, column=1).value,
                    surname=sheet.cell(row=row, column=2).value,
                    name=sheet.cell(row=row, column=3).value,
                    birthday=sheet.cell(row=row, column=4).value
                )
                new_sub.save()
                list.subscribers.add(new_sub)
    except BaseException:
        return False
    return True


